from django.shortcuts import render

# Create your views here.
# apps/paneles/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
# apps/docente/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import date

from .models import (
    Profesor, AsignacionProfesor, Curso, Materia,
    Estudiante, Calificacion, Tarea, Asistencia,
    PeriodoAcademico,
)

# ─────────────────────────────────────────
# PROFESOR
# ─────────────────────────────────────────

def dashboard_profesor(request):
    return render(request, 'paneles/profesor/dashboard_profesor.html')

def cursos_profesor(request):
    return render(request, 'paneles/profesor/cursos_profesor.html')

def calificaciones_profesor(request):
    return render(request, 'paneles/profesor/calificaciones_profesor.html')

def asistencia_profesor(request):
    return render(request, 'paneles/profesor/asistencia_profesor.html')

def historial_asistencia(request):
    return render(request, "paneles/profesor/historial_asistencia.html")

def guardar_asistencia(request):
    return render(request, "paneles/profesor/guardar_asistencia.html")

def configuracion_profesor(request):
    return render(request, 'paneles/profesor/configuracion_profesor.html')


# ─────────────────────────────────────────
# ESTUDIANTE
# ─────────────────────────────────────────


def dashboard_estudiante(request):
    return render(request, 'paneles/estudiante/dashboard_estudiante.html')

def materias_estudiante(request):
    return render(request, 'paneles/estudiante/materias_estudiante.html')


def calificaciones_estudiante(request):
    return render(request, 'paneles/estudiante/calificaciones_estudiante.html')


def asistencia_estudiante(request):
    return render(request, 'paneles/estudiante/asistencia_estudiante.html')


def logros_estudiante(request):
    return render(request, 'paneles/estudiante/logros_estudiante.html')


def configuracion_estudiante(request):
    return render(request, 'paneles/estudiante/configuracion_estudiante.html')

# ─────────────────────────────────────────
# UTILIDAD
# ─────────────────────────────────────────

def _get_profesor(request):
    try:
        return request.user.profesor
    except Profesor.DoesNotExist:
        return None


def _get_periodo_activo():
    return PeriodoAcademico.objects.filter(activo=True).first()


# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────


def dashboard_profesor(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    asignaciones = AsignacionProfesor.objects.filter(
        profesor=profesor
    ).select_related('materia', 'curso')

    tareas = Tarea.objects.filter(asignacion__profesor=profesor)
    mensajes_no_leidos = Mensaje.objects.filter(
        profesor=profesor, leido=False, enviado_por='estudiante'
    ).count()

    # Cursos únicos del docente
    cursos = Curso.objects.filter(
        asignaciones__profesor=profesor
    ).distinct()

    # Tareas recientes
    tareas_recientes = tareas.select_related(
        'asignacion__materia', 'asignacion__curso'
    ).order_by('-fecha_limite')[:5]

    context = {
        'profesor':           profesor,
        'asignaciones':       asignaciones,
        'cursos':             cursos,
        'total_tareas':       tareas.count(),
        'mensajes_no_leidos': mensajes_no_leidos,
        'tareas_recientes':   tareas_recientes,
        'periodo_activo':     _get_periodo_activo(),
    }
    return render(request, 'paneles/profesor/dashboard_profesor.html', context)


# ─────────────────────────────────────────
# CURSOS
# ─────────────────────────────────────────


def cursos_profesor(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    asignaciones = (
        AsignacionProfesor.objects
        .filter(profesor=profesor)
        .select_related('curso', 'materia')
        .prefetch_related('curso__estudiantes')
    )

    context = {
        'profesor':      profesor,
        'asignaciones': asignaciones,
    }
    return render(request, 'paneles/profesor/cursos_profesor.html', context)

# ─────────────────────────────────────────
# CALIFICACIONES (POR PERIODO)
# ─────────────────────────────────────────

def calificaciones_profesor(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    cursos   = Curso.objects.filter(asignaciones__profesor=profesor).distinct()
    materias = Materia.objects.filter(asignaciones__profesor=profesor).distinct()
    periodos = PeriodoAcademico.objects.all().order_by('numero')

    curso_sel   = None
    materia_sel = None
    periodo_sel = None
    estudiantes = []
    calificaciones_map = {}

    # Estadísticas globales
    todas_cal        = Calificacion.objects.filter(asignacion__profesor=profesor)
    total_notas      = todas_cal.count()
    promedios        = [c.promedio for c in todas_cal]
    promedio_general = round(sum(promedios) / len(promedios), 2) if promedios else 0
    en_riesgo        = sum(1 for p in promedios if p < 3.0)
    destacados       = sum(1 for p in promedios if p >= 4.5)

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    periodo_id = request.GET.get('periodo')

    if curso_id and materia_id:
        curso_sel   = get_object_or_404(Curso,   id=curso_id)
        materia_sel = get_object_or_404(Materia, id=materia_id)

        # Periodo seleccionado o activo por defecto
        if periodo_id:
            periodo_sel = get_object_or_404(PeriodoAcademico, id=periodo_id)
        else:
            periodo_sel = _get_periodo_activo()

        asignacion = AsignacionProfesor.objects.filter(
            profesor=profesor, curso=curso_sel, materia=materia_sel
        ).first()

        if asignacion:
            estudiantes = Estudiante.objects.filter(curso=curso_sel)
            cals = Calificacion.objects.filter(
                asignacion=asignacion,
                periodo=periodo_sel,
            )
            calificaciones_map = {c.estudiante_id: c for c in cals}

    context = {
        'profesor':           profesor,
        'cursos':             cursos,
        'materias':           materias,
        'periodos':           periodos,
        'curso_sel':          curso_sel,
        'materia_sel':        materia_sel,
        'periodo_sel':        periodo_sel,
        'estudiantes':        estudiantes,
        'calificaciones_map': calificaciones_map,
        'total_notas':        total_notas,
        'promedio_general':   promedio_general,
        'en_riesgo':          en_riesgo,
        'destacados':         destacados,
    }
    return render(request, 'paneles/profesor/calificaciones_profesor.html', context)



def guardar_nota(request):
    profesor = _get_profesor(request)
    if not profesor:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    curso_id      = request.POST.get('curso_id')
    materia_id    = request.POST.get('materia_id')
    periodo_id    = request.POST.get('periodo_id')

    try:
        tarea_val   = float(request.POST.get('tarea',   0))
        parcial_val = float(request.POST.get('parcial', 0))
        examen_val  = float(request.POST.get('examen',  0))
    except ValueError:
        return JsonResponse({'error': 'Valores inválidos'}, status=400)

    for v in (tarea_val, parcial_val, examen_val):
        if not (0 <= v <= 5):
            return JsonResponse({'error': 'Notas entre 0 y 5'}, status=400)

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)
    asignacion = get_object_or_404(
        AsignacionProfesor,
        profesor=profesor, curso_id=curso_id, materia_id=materia_id,
    )
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id) if periodo_id else None

    cal, _ = Calificacion.objects.get_or_create(
        estudiante=estudiante,
        asignacion=asignacion,
        periodo=periodo,
    )
    cal.tarea   = tarea_val
    cal.parcial = parcial_val
    cal.examen  = examen_val
    cal.save()

    return JsonResponse({'ok': True, 'promedio': cal.promedio})


def eliminar_nota(request):
    profesor = _get_profesor(request)
    if not profesor:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    curso_id      = request.POST.get('curso_id')
    materia_id    = request.POST.get('materia_id')
    periodo_id    = request.POST.get('periodo_id')

    asignacion = get_object_or_404(
        AsignacionProfesor,
        profesor=profesor, curso_id=curso_id, materia_id=materia_id,
    )
    Calificacion.objects.filter(
        estudiante_id=estudiante_id,
        asignacion=asignacion,
        periodo_id=periodo_id or None,
    ).delete()

    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# REPORTES
# ─────────────────────────────────────────

def reporte_notas(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    cursos   = Curso.objects.filter(asignaciones__profesor=profesor).distinct()
    materias = Materia.objects.filter(asignaciones__profesor=profesor).distinct()
    periodos = PeriodoAcademico.objects.all().order_by('numero')

    curso_sel        = None
    materia_sel      = None
    periodo_sel      = None
    calificaciones   = []
    promedio_general = 0

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    periodo_id = request.GET.get('periodo')

    if curso_id and materia_id:
        curso_sel   = get_object_or_404(Curso,   id=curso_id)
        materia_sel = get_object_or_404(Materia, id=materia_id)
        periodo_sel = get_object_or_404(PeriodoAcademico, id=periodo_id) if periodo_id else None

        asignacion = AsignacionProfesor.objects.filter(
            profesor=profesor, curso=curso_sel, materia=materia_sel
        ).first()

        if asignacion:
            qs = Calificacion.objects.filter(
                asignacion=asignacion
            ).select_related('estudiante', 'periodo')
            if periodo_sel:
                qs = qs.filter(periodo=periodo_sel)
            calificaciones = qs
            promedios = [c.promedio for c in calificaciones]
            promedio_general = round(sum(promedios) / len(promedios), 2) if promedios else 0

    context = {
        'profesor': profesor, 'cursos': cursos, 'materias': materias,
        'periodos': periodos, 'curso_sel': curso_sel,
        'materia_sel': materia_sel, 'periodo_sel': periodo_sel,
        'calificaciones': calificaciones, 'promedio_general': promedio_general,
    }
    return render(request, 'paneles/profesor/reporte_notas.html', context)


def exportar_reporte_pdf(request):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    periodo_id = request.GET.get('periodo')

    if not (curso_id and materia_id):
        return redirect('reporte_notas')

    asignacion = get_object_or_404(
        AsignacionProfesor, profesor=profesor, curso_id=curso_id, materia_id=materia_id,
    )
    qs = Calificacion.objects.filter(asignacion=asignacion).select_related('estudiante', 'periodo')
    if periodo_id:
        qs = qs.filter(periodo_id=periodo_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{asignacion.curso}_{asignacion.materia}.pdf"'
    )

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    p.setFont('Helvetica-Bold', 16)
    p.drawString(50, height - 60, f'Reporte — {asignacion.materia} — {asignacion.curso}')
    p.setFont('Helvetica', 11)
    p.drawString(50, height - 80, f'Profesor: {profesor.nombre}')

    y = height - 120
    p.setFont('Helvetica-Bold', 11)
    for col, x in [('Estudiante', 50), ('Periodo', 240), ('Tarea', 310), ('Parcial', 380), ('Examen', 450), ('Promedio', 520)]:
        p.drawString(x, y, col)
    y -= 20

    p.setFont('Helvetica', 10)
    for cal in qs:
        if y < 60:
            p.showPage()
            y = height - 60
        p.drawString(50,  y, cal.estudiante.nombre[:28])
        p.drawString(240, y, cal.periodo.nombre if cal.periodo else '—')
        p.drawString(310, y, str(cal.tarea))
        p.drawString(380, y, str(cal.parcial))
        p.drawString(450, y, str(cal.examen))
        p.drawString(520, y, str(cal.promedio))
        y -= 18

    p.showPage()
    p.save()
    return response



def exportar_reporte_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    periodo_id = request.GET.get('periodo')

    if not (curso_id and materia_id):
        return redirect('reporte_notas')

    asignacion = get_object_or_404(
        AsignacionProfesor, profesor=profesor, curso_id=curso_id, materia_id=materia_id,
    )
    qs = Calificacion.objects.filter(asignacion=asignacion).select_related('estudiante', 'periodo')
    if periodo_id:
        qs = qs.filter(periodo_id=periodo_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte — {asignacion.materia} — {asignacion.curso}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append(['Estudiante', 'Periodo', 'Tarea', 'Parcial', 'Examen', 'Promedio'])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for cal in qs:
        ws.append([
            cal.estudiante.nombre,
            cal.periodo.nombre if cal.periodo else '—',
            float(cal.tarea),
            float(cal.parcial),
            float(cal.examen),
            float(cal.promedio),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{asignacion.curso}_{asignacion.materia}.xlsx"'
    )
    wb.save(response)
    return response


# ─────────────────────────────────────────
# ASISTENCIA (POR ASIGNACIÓN Y FECHA)
# ─────────────────────────────────────────


def asistencia_profesor(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    cursos      = Curso.objects.filter(asignaciones__profesor=profesor  ).distinct()
    asignaciones_profesor = AsignacionProfesor.objects.filter(
        profesor=profesor
    ).select_related('materia', 'curso')

    fecha_hoy = timezone.now().date()
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        try:
            fecha_sel = date.fromisoformat(fecha_str)
            if fecha_sel > fecha_hoy:
                fecha_sel = fecha_hoy
        except ValueError:
            fecha_sel = fecha_hoy
    else:
        fecha_sel = fecha_hoy

    curso_sel      = None
    asignacion_sel = None
    estudiantes    = []
    asistencias_map = {}

    curso_id      = request.GET.get('curso')
    asignacion_id = request.GET.get('asignacion')

    if curso_id:
        curso_sel = get_object_or_404(Curso, id=curso_id)
        # Asignaciones del profesor en ese curso
        asignaciones_curso = asignaciones_profesor.filter(curso=curso_sel)

        if asignacion_id:
            asignacion_sel = get_object_or_404(
                AsignacionProfesor, id=asignacion_id, profesor=profesor, curso=curso_sel
            )
            estudiantes = Estudiante.objects.filter(curso=curso_sel)
            asis = Asistencia.objects.filter(
                estudiante__in=estudiantes,
                asignacion=asignacion_sel,
                fecha=fecha_sel,
            )
            asistencias_map = {a.estudiante_id: a.estado for a in asis}
    else:
        asignaciones_curso = asignaciones_profesor.none()

    total     = len(estudiantes)
    presentes = sum(1 for v in asistencias_map.values() if v == 'presente')
    tardanzas = sum(1 for v in asistencias_map.values() if v == 'tardanza')
    ausentes  = sum(1 for v in asistencias_map.values() if v == 'ausente')

    context = {
        'profesor':           profesor,
        'cursos':             cursos,
        'asignaciones_curso': asignaciones_curso if curso_id else [],
        'curso_sel':          curso_sel,
        'asignacion_sel':     asignacion_sel,
        'estudiantes':        estudiantes,
        'asistencias_map':    asistencias_map,
        'fecha_hoy':          fecha_hoy,
        'fecha_sel':          fecha_sel,
        'total':              total,
        'presentes':          presentes,
        'tardanzas':          tardanzas,
        'ausentes':           ausentes,
    }
    return render(request, 'profesor/asistencia_profesor.html', context)


def guardar_asistencia(request):
    profesor = _get_profesor(request)
    if not profesor:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    fecha         = request.POST.get('fecha')
    estado        = request.POST.get('estado', 'presente')
    asignacion_id = request.POST.get('asignacion_id')

    if not (estudiante_id and fecha and asignacion_id):
        return JsonResponse({'error': 'Faltan datos'}, status=400)

    if estado not in {'presente', 'ausente', 'tardanza'}:
        return JsonResponse({'error': 'Estado no válido'}, status=400)

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)
    asignacion = get_object_or_404(
        AsignacionProfesor, id=asignacion_id, profesor=profesor, curso=estudiante.curso
    )

    Asistencia.objects.update_or_create(
        estudiante=estudiante,
        asignacion=asignacion,
        fecha=fecha,
        defaults={'estado': estado},
    )
    return JsonResponse({'ok': True})


def historial_asistencia(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    cursos   = Curso.objects.filter(asignaciones__profesor=profesor).distinct()
    hoy      = timezone.now().date()

    curso_id      = request.GET.get('curso')
    asignacion_id = request.GET.get('asignacion')
    fecha_str     = request.GET.get('fecha')

    if fecha_str:
        try:
            fecha_sel = date.fromisoformat(fecha_str)
        except ValueError:
            fecha_sel = hoy
    else:
        fecha_sel = hoy

    curso_sel          = None
    asignacion_sel     = None
    asignaciones_curso = []
    filas              = []

    if curso_id:
        curso_sel = get_object_or_404(Curso, id=curso_id)
        asignaciones_curso = AsignacionProfesor.objects.filter(
            profesor=profesor, curso=curso_sel
        ).select_related('materia')

        if asignacion_id:
            asignacion_sel = get_object_or_404(
                AsignacionProfesor, id=asignacion_id, profesor=profesor, curso=curso_sel
            )
            estudiantes = Estudiante.objects.filter(curso=curso_sel)
            asis = Asistencia.objects.filter(
                estudiante__in=estudiantes,
                asignacion=asignacion_sel,
                fecha=fecha_sel,
            )
            asistencia_map = {a.estudiante_id: a for a in asis}
            filas = [
                {'estudiante': est, 'asistencia': asistencia_map.get(est.id)}
                for est in estudiantes
            ]

    context = {
        'profesor':           profesor,
        'cursos':             cursos,
        'asignaciones_curso': asignaciones_curso,
        'curso_sel':          curso_sel,
        'asignacion_sel':     asignacion_sel,
        'fecha_sel':          fecha_sel,
        'filas':              filas,
        'hoy':                hoy,
    }
    return render(request, 'paneles/profesor/historial_asistencia.html', context)


def eliminar_asistencia(request):
    profesor = _get_profesor(request)
    if not profesor:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    fecha         = request.POST.get('fecha')
    asignacion_id = request.POST.get('asignacion_id')

    if not (estudiante_id and fecha):
        return JsonResponse({'error': 'Faltan datos'}, status=400)

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)

    es_del_profesor = AsignacionProfesor.objects.filter(
        profesor=profesor, curso=estudiante.curso
    ).exists()
    if not es_del_profesor:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    qs = Asistencia.objects.filter(estudiante=estudiante, fecha=fecha)
    if asignacion_id:
        qs = qs.filter(asignacion_id=asignacion_id)
    qs.delete()

    return JsonResponse({'ok': True})



# ─────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────

def configuracion_profesor(request):
    profesor = _get_profesor(request)
    if not profesor:
        return redirect('iniciar_sesion')

    if request.method == 'POST':
        profesor.nombre   = request.POST.get('nombre',   profesor.nombre).strip()
        profesor.correo   = request.POST.get('correo',   profesor.correo).strip()
        profesor.telefono = request.POST.get('telefono', profesor.telefono).strip()
        profesor.save()
        return redirect('configuracion_profesor')

    context = {'profesor': profesor}
    return render(request, 'paneles/profesor/configuracion_profesor.html', context)